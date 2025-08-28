import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl 
from openpyxl.styles import Alignment 
import shutil 
import os
from datetime import date
import sys

#https://www.youtube.com/watch?v=p3tSLatmGvU&t=660s

#https://stackoverflow.com/questions/31836104/pyinstaller-and-onefile-how-to-include-an-image-in-the-exe-file
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS2
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

def upload_file():
    file_path = filedialog.askopenfilename(
        title="Select an Excel file",
        filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*"))
    )

    if file_path: 
        # destination_folder = os.path.join("C:","EazyAmortization")
        destination_folder = resource_path("EazyAmortization")
        
        os.makedirs(destination_folder, exist_ok=True)
        
        file_name = os.path.basename(file_path)
        
        destination_path = os.path.join(destination_folder, file_name)
        
        try:
            shutil.copy(file_path, destination_path)
            show_alert("Success", f"File '{file_name}' uploaded successfully to '{destination_folder}'!")
            update_drop()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to upload file: {e}")
        

def get_drop_items():
    # entries = os.listdir(os.path.join("C:","EazyAmortization"))
    entries = os.listdir(resource_path("EazyAmortization"))
    for i in range(len(entries)):
        entries[i] = entries[i][:entries[i].index(".")]
    return entries

def update_drop():
    global drop_widget, drop_opt
    entries = get_drop_items()
    drop_widget.destroy()
    drop_opt.set(entries[0])
    drop_widget = tk.OptionMenu(root, drop_opt, *entries)
    drop_widget.place(x=175,y=100)

def button_pressed():
    if not date_validate():
        return
    date_ls = date_var.get().split("-")
    try:
        date_date = date(int(date_ls[2]),int(date_ls[0]),int(date_ls[1]))
    except:
        show_alert("Date Formating Error","Date doesnt exist make sure it's MM-DD-YYYY")
        return
    write_to_excel(drop_opt.get(),date_date,paid_var.get())
    
def only_num(new_value):
    n_val = new_value
    if len(n_val) > 10:
        return False 
    if new_value == "":
        return True
    try:
        if "-" in new_value:
            n_val = n_val.replace("-","")
        if "." in new_value:
            n_val = n_val.replace(".","")  
        int(n_val)
        return True
    except ValueError:
        return False

def date_validate():

    date_ls = date_var.get().split("-")
    if date_ls == [""]:
        return False
    if len(date_ls) != 3:
        show_alert("Date Format Missing Hyphon","You are missing a hyphon -")
        return False
    elif len(date_ls[0]) != 2 or len(date_ls[1]) != 2 or len(date_ls[2]) != 4:
        show_alert("Date Format Error","Must write Date as MM-DD-YYYY")
        return False
    future = False
    if int(date_ls[2]) > date.today().year:
        future = True
    if int(date_ls[2]) == date.today().year and int(date_ls[0]) > date.today().month:
        future = True
    if int(date_ls[2]) == date.today().year and int(date_ls[0]) == date.today().month and int(date_ls[1]) > date.today().day:
        future = True
    if future:
        if not show_warning("Future Payment Warning",f"{date_var.get()} is a future payment"):
            return False
    return True

def show_alert(title, message):
    messagebox.showinfo(title, message)

def show_warning(title,message):
    return messagebox.askokcancel(title, message)

def write_to_excel(client,date_date,paid):
    path = resource_path(f"EazyAmortization\\{client}.xlsx")
    workbook = openpyxl.load_workbook(path)
    sheet = workbook["Amortization Fixed"]

    date_book = workbook["Cover"]
    date_val = date_book["D4"]
    row = months_apart(date_val.value,date_date)

    date_cell = sheet[f"C{row}"]
    paid_cell = sheet[f"N{row}"]

    if date_cell.value != None or paid_cell.value != None:
        if not show_warning("Replace Payment Info",f"Replace payment of {paid_cell.value} on {date_cell.value.strftime(r"%m/%d/%Y")} "
                            f"with payment of {paid} on {date_date.strftime(r"%m/%d/%Y")}?"):
            return
    date_cell.number_format = 'm/d/YYYY'
    date_cell.value = date_date
    date_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    paid_cell.number_format = '$#,##0.00'
    paid_cell.value = paid
    paid_cell.alignment = Alignment(horizontal="center",vertical="center")
    
    try:
        workbook.save(path)
    except PermissionError:
        show_alert("File Open","You must close the file before saving")
        return
    show_alert("Payment Made",f"{paid} paid on {date_date.strftime(r"%m/%d/%Y")}")


def months_apart(dt1,dt2):
    dy = dt2.year - dt1.year
    dm = dt2.month - dt1.month
    return dy*12+dm+2

def get_date_cell(client):
    date_val = get_last_paid(client,False)
    date_date = date_val.value.date()
    d = date(date_date.year, date_date.month +1,1)
    return d.strftime(r"%m-%d-%Y")

def get_expected_paid_cell(client):
    expected_paid = get_last_paid(client,True)
    if expected_paid.value == None:
        return 0
    return expected_paid.value
    
def get_last_paid(client,boolean):
    path = resource_path(f"EazyAmortization\\{client}.xlsx")
    try:
        workbook = openpyxl.load_workbook(path)
    except PermissionError:
        show_alert("Permission Denied Error",f"{client} is open.\nClose the file to continue")

    sheet = workbook["Amortization Fixed"]
    # will return payment if true or date if false
    col = "N" if boolean else "C"
    expected_paid = sheet[f"{col}2"]
    for i in range(sheet.max_row,1,-1):
        # this is to check for the last payment
        if sheet[f"N{i}"].value is not None:
            expected_paid = sheet[f"{col}{i}"]
            break
    if not boolean:
        if expected_paid.value == None:
            return workbook["Cover"]["D4"]    
    return expected_paid

def switch_vars(self):
    date_var.set(get_date_cell(self))
    paid_var.set(get_expected_paid_cell(self))

def open_pressed():
    os.startfile(resource_path(f"EazyAmortization\\{drop_opt.get()}.xlsx"))
def delete_pressed():
    if show_warning("Delete File",f"Are you sure you want to delete {drop_opt.get()}") and os.path.exists(f"EazyAmortization\\{drop_opt.get()}.xlsx"):
        os.remove(resource_path(f"EazyAmortization\\{drop_opt.get()}.xlsx"))
        update_drop()
root = tk.Tk()
root.title("Eazy Amortization")
root.geometry("500x350") 
root.resizable(False, False) 


upload_button = tk.Button(root, text="Upload Excel File", command=upload_file,
                          bg="#4CAF50", fg="white", font=("Arial", 12, "bold"),
                          padx=1, pady=5, relief="raised", bd=3,
                          activebackground="#45a049", activeforeground="white",cursor="hand2")
upload_button.pack(pady=10)


instruction_label = tk.Label(root, text="Click the button to select and upload an Excel file.",font=("Arial", 10), fg="gray")
instruction_label.pack(pady=1)

drop_opt = tk.StringVar(root)
drop_items = get_drop_items()

drop_opt.set(drop_items[0])
drop_widget = tk.OptionMenu(root, drop_opt, *drop_items,command=switch_vars)

drop_widget.config(cursor="hand2")
drop_widget.place(x=175,y=100)

input_frame = tk.LabelFrame(root, text="Cell Data Input", padx=10, pady=10)
input_frame.pack(pady=50, padx=10, fill="x")

vd = root.register(only_num)

date_var = tk.StringVar()
date_var.set(get_date_cell(drop_opt.get()))

date_label = tk.Label(input_frame, text="Enter Date (MM-DD-YYYY):")
date_label.grid(row=1, column=0, sticky="w", pady=5, padx=5)
date_entry = tk.Entry(input_frame, width=15,textvariable=date_var,validate="key",validatecommand=(vd,"%P"))
date_entry.grid(row=1, column=1, pady=5, padx=5)

paid_var = tk.DoubleVar()
paid_var.set(get_expected_paid_cell(drop_opt.get()))

paid_label = tk.Label(input_frame, text="Enter Amount:")
paid_label.grid(row=1, column=2, sticky="w", pady=5, padx=5)
paid_entry = tk.Entry(input_frame, width=15,validate="key",validatecommand=(vd,"%P"),textvariable=paid_var)
paid_entry.grid(row=1, column=3, pady=5, padx=5)



format_button = tk.Button(input_frame, text="Enter",padx=20,command=button_pressed,cursor="hand2")
format_button.grid(pady=10, row=2, column=1)

open_button = tk.Button(input_frame, text="Open",padx=20,command=open_pressed,cursor="hand2")
open_button.grid(pady=10, row=2, column=2)

delete_button = tk.Button(input_frame, text="Delete",padx=20,command=delete_pressed,cursor="hand2")
delete_button.grid(pady=10,row=2,column=3)
root.mainloop()